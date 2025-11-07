"""
Hybrid Triage Financial Data Cleaner
=====================================
This script uses a three-step triage process to map messy cost descriptions 
to clean actual GL Heads:
1. Exact Match - For perfect matches
2. Fuzzy Match - For typos and minor variations
3. AI Escalation - For complex cases requiring context understanding

Author: Sambhav Foundation
Date: November 2025
"""

import pandas as pd
import json
from thefuzz import process
import ollama
import numpy as np
from tqdm import tqdm
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


def load_data(file_path='Data/Test.xlsx'):
    """
    Load the input Excel file, reading all sheets and separating rules/messy.
    
    Returns:
        tuple: (df_rules, df_messy, all_sheets)
    """
    print(f"Loading all sheets from {file_path}...")
    
    # Read all sheets into a dictionary of DataFrames
    # header=0 ensures the first row is read as the header for all sheets
    all_sheets = pd.read_excel(file_path, sheet_name=None, header=0)
    
    # Get the specific sheets we need
    df_rules = all_sheets['rules'].copy()
    df_messy = all_sheets['messy'].copy()
    
    # Normalize column names for our key dataframes
    df_rules.columns = df_rules.columns.str.strip()
    df_messy.columns = df_messy.columns.str.strip()
    
    # Rename columns to match expected names
    column_mapping_rules = {
        'SF Budget/Procurement Head': 'procurement head',
        'SF Cost/GL Head': 'actual GL Head',
        'Rules': 'Rules_and_Examples'
    }
    df_rules = df_rules.rename(columns=column_mapping_rules)
    
    print(f"Loaded {len(all_sheets)} total sheets.")
    print(f"Found {len(df_rules)} rules and {len(df_messy)} messy entries.")
    return df_rules, df_messy, all_sheets


def initialize_output_columns(df_messy):
    """
    Add three new columns to the messy dataframe for predictions.
    
    Args:
        df_messy: The dataframe with messy cost data
        
    Returns:
        pd.DataFrame: Modified dataframe with new columns
    """
    df_messy['Predicted cost head'] = None
    df_messy['Prediction Confidence'] = None
    df_messy['AI Confidence'] = None
    
    return df_messy


def get_valid_gl_heads(df_rules):
    """
    Extract the list of valid actual GL Heads from the rules dataframe.
    
    Args:
        df_rules: The rules dataframe
        
    Returns:
        list: List of valid actual GL Head strings
    """
    valid_heads = df_rules['actual GL Head'].dropna().unique().tolist()
    print(f"Found {len(valid_heads)} valid actual GL Heads.")
    return valid_heads


def step1_exact_match(df_messy, valid_gl_heads):
    """
    Step 1: Exact Match
    Find rows where the Cost head exactly matches an actual GL Head.
    
    Args:
        df_messy: The messy dataframe
        valid_gl_heads: List of valid GL heads
        
    Returns:
        int: Number of exact matches found
    """
    print("\n--- Step 1: Exact Match ---")
    
    # Convert valid_gl_heads to a set for faster lookup
    valid_set = set(valid_gl_heads)
    exact_matches = 0
    
    for idx, row in df_messy.iterrows():
        # Skip if already processed
        if pd.notna(row['Predicted cost head']):
            continue
            
        cost_head = row['Cost head']
        
        # Check for exact match
        if cost_head in valid_set:
            df_messy.at[idx, 'Predicted cost head'] = cost_head
            df_messy.at[idx, 'Prediction Confidence'] = 1.0
            exact_matches += 1
    
    print(f"Found {exact_matches} exact matches.")
    return exact_matches


def step2_fuzzy_match(df_messy, valid_gl_heads, threshold=90):
    """
    Step 2: Fuzzy Match
    Use fuzzy string matching to find typos and minor variations.
    
    Args:
        df_messy: The messy dataframe
        valid_gl_heads: List of valid GL heads
        threshold: Minimum fuzzy match score (0-100)
        
    Returns:
        int: Number of fuzzy matches found
    """
    print("\n--- Step 2: Fuzzy Match ---")
    fuzzy_matches = 0
    
    for idx, row in df_messy.iterrows():
        # Skip if already processed
        if pd.notna(row['Predicted cost head']):
            continue
            
        cost_head = row['Cost head']
        
        # Find the best fuzzy match
        best_match = process.extractOne(cost_head, valid_gl_heads)
        
        if best_match and best_match[1] > threshold:
            matched_string, score = best_match[0], best_match[1]
            df_messy.at[idx, 'Predicted cost head'] = matched_string
            df_messy.at[idx, 'Prediction Confidence'] = score / 100.0
            fuzzy_matches += 1
    
    print(f"Found {fuzzy_matches} fuzzy matches (threshold > {threshold}%).")
    return fuzzy_matches


def get_ai_prediction(messy_string, system_prompt, valid_gl_heads):
    """
    Use Ollama AI to predict the correct GL Head based on context.
    
    Args:
        messy_string: The messy cost description
        system_prompt: The pre-built system prompt containing all rules
        valid_gl_heads: List of valid GL heads (for validation)
        
    Returns:
        tuple: (prediction, ai_confidence) or (None, 0) if failed
    """
    
    try:
        # Call Ollama with a separate system and user prompt
        # Ollama will cache the system_prompt, making subsequent calls
        # much faster.
        response = ollama.chat(
            model='qwen3:4b',
            messages=[
                {
                    'role': 'system',
                    'content': system_prompt
                },
                {
                    'role': 'user',
                    'content': f"Now, analyze the following messy Cost head: '{messy_string}'"
                }
            ],
            format='json',
            think=False  # Disable thinking process for faster inference
        )
        
        # Parse the JSON response
        result = json.loads(response['message']['content'])
        prediction = result.get('prediction')
        ai_confidence = result.get('ai_confidence', 0)
        
        # Validate the prediction is in the valid list
        if prediction not in valid_gl_heads:
            print(f"  Warning: AI returned invalid GL Head '{prediction}'. Using best fuzzy match.")
            # Find closest valid match
            best_match = process.extractOne(prediction, valid_gl_heads)
            if best_match:
                prediction = best_match[0]
        
        return prediction, ai_confidence
        
    except Exception as e:
        print(f"  Error calling AI: {e}")
        return None, 0


def step3_ai_escalation(df_messy, df_rules, valid_gl_heads):
    """
    Step 3: AI Escalation
    Use Ollama AI for remaining unprocessed rows.
    
    Args:
        df_messy: The messy dataframe
        df_rules: The rules dataframe
        valid_gl_heads: List of valid GL heads
        
    Returns:
        int: Number of AI predictions made
    """
    print("\n--- Step 3: AI Escalation ---")
    ai_predictions = 0
    
    # --- PROMPT BUILDING (MOVED OUTSIDE LOOP) ---
    print("  Building AI system prompt...")
    valid_heads_list = "\n".join([f"- {head}" for head in valid_gl_heads])
    
    rules_context = []
    for _, rule in df_rules.iterrows():
        rules_context.append(
            f"Category: {rule['procurement head']}\n"
            f"  Actual GL Head: {rule['actual GL Head']}\n"
            f"  Rules/Examples: {rule['Rules_and_Examples']}"
        )
    rules_str = "\n\n".join(rules_context)
    
    system_prompt = f"""You are an expert financial analyst. Your task is to map a 'messy' cost description to a single, official 'actual GL Head'.

You MUST choose one and only one of the following valid 'actual GL Head' options:
{valid_heads_list}

To help you, here is the full rulebook, including procurement categories and specific examples:

{rules_str}

You MUST respond only with a valid JSON object. Do not add any other text. The JSON must have two keys:
- "prediction": the string of the 'actual GL Head' you chose (must be from the list above)
- "ai_confidence": a number from 1 to 10 rating your confidence (10 = very confident, 1 = just guessing)

Example response format:
{{"prediction": "Consumables - Printing", "ai_confidence": 8}}
"""
    # --- END OF PROMPT BUILDING ---
    
    # Get all rows that need AI processing
    unprocessed_indices = df_messy[pd.isna(df_messy['Predicted cost head'])].index
    print(f"  Processing {len(unprocessed_indices)} entries with AI...")
    
    # Use tqdm for a progress bar
    for idx in tqdm(unprocessed_indices, desc="AI Predictions"):
        cost_head = df_messy.at[idx, 'Cost head']
        
        # Get AI prediction
        prediction, ai_confidence = get_ai_prediction(cost_head, system_prompt, valid_gl_heads)
        
        if prediction:
            df_messy.at[idx, 'Predicted cost head'] = prediction
            df_messy.at[idx, 'AI Confidence'] = ai_confidence
            df_messy.at[idx, 'Prediction Confidence'] = ai_confidence / 10.0
            ai_predictions += 1
        else:
            print(f"    → Failed to get prediction for '{cost_head}'")
    
    print(f"Completed {ai_predictions} AI predictions.")
    return ai_predictions


def apply_excel_formatting(df, writer, sheet_name='messy'):
    """
    Apply column widths and conditional formatting to the 'messy' sheet.
    
    Args:
        df: The dataframe (used to find column names)
        writer: The Pandas ExcelWriter object
        sheet_name: The name of the sheet to format
    """
    print(f"  Applying formatting to '{sheet_name}' sheet...")
    try:
        # Get the openpyxl workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # 1. Set Column Widths
        # Set all columns to a width of 30 (~210 pixels)
        for i, col in enumerate(df.columns, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = 30

        # 2. Add Conditional Formatting (Gradients)
        
        # Find the column letters for our confidence scores
        confidence_col_letter = None
        ai_confidence_col_letter = None
        
        for i, col_name in enumerate(df.columns, 1):
            if col_name == 'Prediction Confidence':
                confidence_col_letter = get_column_letter(i)
            if col_name == 'AI Confidence':
                ai_confidence_col_letter = get_column_letter(i)

        # Define the data range (e.g., "F2:F28")
        # We start at row 2 (header=0, data=1 + 1-based index = 2)
        max_row = len(df) + 1

        # Apply 3-color scale to 'Prediction Confidence' (0.0 to 1.0)
        if confidence_col_letter:
            col_range = f"{confidence_col_letter}2:{confidence_col_letter}{max_row}"
            # Red (min) -> Yellow (50%) -> Green (max)
            rule = ColorScaleRule(
                start_type='min', start_color='FFB0B0',  # Light Red
                mid_type='percentile', mid_value=50, mid_color='FFFFB0',  # Light Yellow
                end_type='max', end_color='B0FFB0'  # Light Green
            )
            worksheet.conditional_formatting.add(col_range, rule)

        # Apply 3-color scale to 'AI Confidence' (1 to 10)
        if ai_confidence_col_letter:
            col_range = f"{ai_confidence_col_letter}2:{ai_confidence_col_letter}{max_row}"
            # Red (1) -> Yellow (5) -> Green (10)
            rule = ColorScaleRule(
                start_type='num', start_value=1, start_color='FFB0B0',  # Light Red
                mid_type='num', mid_value=5, mid_color='FFFFB0',  # Light Yellow
                end_type='num', end_value=10, end_color='B0FFB0'  # Light Green
            )
            worksheet.conditional_formatting.add(col_range, rule)
            
        print("  Formatting applied successfully.")

    except Exception as e:
        print(f"  Warning: Could not apply Excel formatting. Error: {e}")


def save_results(df_messy, all_sheets, output_file='Data/result.xlsx'):
    """
    Save the processed dataframe and all other original sheets to a new file.
    
    Args:
        df_messy: The processed dataframe
        all_sheets: The dictionary of all original dataframes
        output_file: Path to the output file
    """
    print(f"\nSaving results to {output_file}...")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, df in all_sheets.items():
            if sheet_name == 'messy':
                # Write our *modified* messy sheet
                df_messy.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                # Write all other sheets back as they were
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # --- NEW CODE ---
        # Now that the 'messy' sheet is written, apply formatting
        # We pass df_messy so it can get column names, and the writer
        # to get the openpyxl worksheet object.
        apply_excel_formatting(df_messy, writer, sheet_name='messy')
        # --- END NEW CODE ---
    
    print(f"Processing complete. All sheets saved to {output_file}")


def print_summary(df_messy):
    """
    Print a summary of the processing results.
    
    Args:
        df_messy: The processed dataframe
    """
    print("\n" + "="*60)
    print("PROCESSING SUMMARY")
    print("="*60)
    
    total = len(df_messy)
    processed = df_messy['Predicted cost head'].notna().sum()
    unprocessed = total - processed
    
    exact = (df_messy['Prediction Confidence'] == 1.0).sum()
    fuzzy = ((df_messy['Prediction Confidence'] > 0) & 
             (df_messy['Prediction Confidence'] < 1.0) & 
             (df_messy['AI Confidence'].isna())).sum()
    ai = df_messy['AI Confidence'].notna().sum()
    
    print(f"Total entries: {total}")
    print(f"Processed: {processed} ({processed/total*100:.1f}%)")
    print(f"  - Exact matches: {exact}")
    print(f"  - Fuzzy matches: {fuzzy}")
    print(f"  - AI predictions: {ai}")
    print(f"Unprocessed: {unprocessed}")
    
    if processed > 0:
        avg_confidence = df_messy['Prediction Confidence'].mean()
        print(f"\nAverage confidence: {avg_confidence:.2f}")
    
    print("="*60)


def main():
    """
    Main execution function.
    """
    print("="*60)
    print("Hybrid Triage Financial Data Cleaner")
    print("="*60)
    
    # Load data (now returns all_sheets as well)
    df_rules, df_messy, all_sheets = load_data('Data/Test.xlsx')
    
    # Initialize output columns
    df_messy = initialize_output_columns(df_messy)
    
    # Get valid GL heads
    valid_gl_heads = get_valid_gl_heads(df_rules)
    
    # Execute the three-step triage process
    step1_exact_match(df_messy, valid_gl_heads)
    step2_fuzzy_match(df_messy, valid_gl_heads, threshold=90)
    step3_ai_escalation(df_messy, df_rules, valid_gl_heads)
    
    # Print summary
    print_summary(df_messy)
    
    # Save results (now passes all_sheets)
    save_results(df_messy, all_sheets, 'Data/result.xlsx')


if __name__ == "__main__":
    main()
